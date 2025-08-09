from openpyxl import load_workbook

planilha_vendas = load_workbook('vendas_carros.xlsx')
pagina_vendas = planilha_vendas['Sheet1']

for linha in pagina_vendas.iter_rows(values_only=True):
    print(linha)
